import json
from openpyxl import Workbook


def txt_to_excel(path):
    file = open(path)
    ob_json = json.load(file)
    print(ob_json)
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    for i in range(1, len(ob_json)+1):
        worksheet.cell(row=i, column=1).value = i
        for j in range(0,len(ob_json[str(i)])):
            worksheet.cell(row=i, column=j+2).value = ob_json[str(i)][j]
    workbook.save(filename=path.replace('txt', 'xls'))


if __name__ == '__main__':
    txt_to_excel('student.txt')